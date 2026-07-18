import os
import base64
import logging
from pathlib import Path


def is_configured() -> bool:
    return bool(os.getenv("JIRA_URL") and os.getenv("JIRA_EMAIL") and os.getenv("JIRA_TOKEN"))


def setup_credentials(url: str, email: str, token: str) -> None:
    """Guarda credenciales Jira en ~/.mycontext/.env y actualiza el proceso actual."""
    from pathlib import Path
    env_path = Path.home() / ".mycontext" / ".env"

    existing = env_path.read_text(encoding="utf-8") if env_path.exists() else ""
    lines = [ln for ln in existing.splitlines() if ln.strip()]

    new_vars = {"JIRA_URL": url.rstrip("/"), "JIRA_EMAIL": email, "JIRA_TOKEN": token}
    updated: set[str] = set()
    result: list[str] = []

    for line in lines:
        key = line.split("=", 1)[0] if "=" in line else ""
        if key in new_vars:
            result.append(f"{key}={new_vars[key]}")
            updated.add(key)
        else:
            result.append(line)

    for key, val in new_vars.items():
        if key not in updated:
            result.append(f"{key}={val}")

    env_path.write_text("\n".join(result) + "\n", encoding="utf-8")

    for key, val in new_vars.items():
        os.environ[key] = val


def _headers() -> dict:
    creds = base64.b64encode(
        f"{os.getenv('JIRA_EMAIL')}:{os.getenv('JIRA_TOKEN')}".encode()
    ).decode()
    return {"Authorization": f"Basic {creds}", "Accept": "application/json"}


def _adf_to_text(node) -> str:
    """Convierte Atlassian Document Format (ADF) a texto plano."""
    if node is None:
        return ""
    if isinstance(node, str):
        return node
    if isinstance(node, list):
        return "".join(_adf_to_text(n) for n in node)
    t = node.get("type", "")
    content = node.get("content", [])
    if t == "text":
        return node.get("text", "")
    if t in ("paragraph", "heading"):
        inner = _adf_to_text(content)
        return (inner + "\n") if inner.strip() else ""
    if t == "bulletList":
        return "".join(
            f"• {_adf_to_text(i.get('content', [])).strip()}\n"
            for i in content
        )
    if t == "orderedList":
        return "".join(
            f"{idx + 1}. {_adf_to_text(i.get('content', [])).strip()}\n"
            for idx, i in enumerate(content)
        )
    if t == "hardBreak":
        return "\n"
    return _adf_to_text(content)


def fetch_issue(ticket_id: str) -> dict | None:
    import httpx
    url = f"{os.getenv('JIRA_URL')}/rest/api/3/issue/{ticket_id}"
    try:
        resp = httpx.get(url, headers=_headers(), timeout=10)
        if resp.status_code != 200:
            logging.warning("jira.fetch_issue %s — HTTP %d", ticket_id, resp.status_code)
            return None
        fields = resp.json().get("fields", {})

        desc_raw = fields.get("description")
        description = (
            _adf_to_text(desc_raw).strip()
            if isinstance(desc_raw, dict)
            else (desc_raw or "")
        )

        return {
            "id": ticket_id.upper(),
            "summary": fields.get("summary", ""),
            "description": description,
            "priority": (fields.get("priority") or {}).get("name", ""),
            "status": (fields.get("status") or {}).get("name", ""),
            "reporter": (fields.get("reporter") or {}).get("displayName", ""),
            "assignee": (fields.get("assignee") or {}).get("displayName", ""),
            "attachments": fields.get("attachment", []),
        }
    except Exception as e:
        logging.warning("jira.fetch_issue %s — %s", ticket_id, e)
        return None


def fetch_my_issues() -> dict:
    """Trae issues asignados al usuario actual, agrupados por estado."""
    import httpx
    url = f"{os.getenv('JIRA_URL')}/rest/api/3/search"
    jql = (
        "assignee = currentUser() "
        "AND statusCategory in (indeterminate, new) "
        "ORDER BY priority DESC, updated DESC"
    )
    try:
        resp = httpx.get(
            url, headers=_headers(),
            params={"jql": jql, "maxResults": 20, "fields": "summary,status,priority"},
            timeout=10,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"HTTP {resp.status_code} — {resp.text[:120]}")

        _HIGH = {"high", "highest", "critical", "alta", "crítica", "critica"}
        result: dict = {"en_curso": [], "alta_prioridad": [], "reabiertos": []}

        for issue in resp.json().get("issues", []):
            f = issue.get("fields", {})
            item = {
                "id": issue.get("key", ""),
                "summary": f.get("summary", ""),
                "status": (f.get("status") or {}).get("name", ""),
                "priority": (f.get("priority") or {}).get("name", ""),
            }
            status_cat = (f.get("status") or {}).get("statusCategory", {}).get("key", "")
            status_lower = item["status"].lower()
            prio_lower = item["priority"].lower()

            if "reabiert" in status_lower or "reopened" in status_lower:
                result["reabiertos"].append(item)
            elif status_cat == "indeterminate":
                result["en_curso"].append(item)
            elif prio_lower in _HIGH:
                result["alta_prioridad"].append(item)

        return result
    except Exception as e:
        logging.warning("jira.fetch_my_issues — %s", e)
        return {}


def download_image_attachments(attachments: list) -> list[str]:
    """Descarga adjuntos de imagen al directorio de evidencias. Retorna rutas locales."""
    import httpx
    _IMAGE = {"image/png", "image/jpeg", "image/gif", "image/webp"}
    folder = Path.home() / ".mycontext" / "evidencias"
    paths = []
    for att in attachments:
        if att.get("mimeType", "") not in _IMAGE:
            continue
        try:
            resp = httpx.get(att["content"], headers=_headers(), timeout=30, follow_redirects=True)
            if resp.status_code == 200:
                dest = folder / att.get("filename", "jira_attachment.png")
                dest.write_bytes(resp.content)
                paths.append(str(dest))
        except Exception as e:
            logging.warning("jira.download_image %s — %s", att.get("filename"), e)
    return paths


_EXCEL_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


def download_excel_attachments(attachments: list) -> list[str]:
    """Descarga adjuntos Excel al directorio de evidencias. Retorna rutas locales."""
    import httpx
    folder = Path.home() / ".mycontext" / "evidencias"
    paths = []
    for att in attachments:
        if att.get("mimeType", "") not in _EXCEL_MIME:
            continue
        try:
            resp = httpx.get(att["content"], headers=_headers(), timeout=30, follow_redirects=True)
            if resp.status_code == 200:
                dest = folder / att.get("filename", "jira_attachment.xlsx")
                dest.write_bytes(resp.content)
                paths.append(str(dest))
        except Exception as e:
            logging.warning("jira.download_excel %s — %s", att.get("filename"), e)
    return paths


_EXCEL_MAX_ROWS = 200


def excel_to_text(path: str) -> str:
    """Convierte un archivo .xlsx a texto markdown con tablas por hoja."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sections = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            truncated = False
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([str(c) if c is not None else "" for c in row])
                    if len(rows) >= _EXCEL_MAX_ROWS + 1:
                        truncated = True
                        break
            if not rows:
                continue
            header = "| " + " | ".join(rows[0]) + " |"
            separator = "| " + " | ".join(["---"] * len(rows[0])) + " |"
            body = "\n".join("| " + " | ".join(r) + " |" for r in rows[1:])
            note = f"\n\n*(truncado en {_EXCEL_MAX_ROWS} filas)*" if truncated else ""
            sections.append(f"### Hoja: {sheet_name}\n\n{header}\n{separator}\n{body}{note}")
        wb.close()
        return "\n\n".join(sections) if sections else "(sin datos)"
    except Exception as e:
        logging.warning("jira.excel_to_text %s — %s", path, e)
        return f"(no se pudo leer: {e})"
